import cv2
import numpy as np
import mss
import openpyxl
from paddleocr import PaddleOCR

ocr = PaddleOCR(use_angle_cls=False, lang='en', use_gpu=True)

def preprocess_image(image):
    resized_img = cv2.resize(image, None, fx=1.7, fy=1.7, interpolation=cv2.INTER_CUBIC)
    gray_img = cv2.cvtColor(resized_img, cv2.COLOR_BGR2GRAY)
    _, thresh_img = cv2.threshold(gray_img, 135, 255, cv2.THRESH_BINARY_INV)
    return thresh_img

def load_templates():
    weapon_templates = {
        'Ares': cv2.imread('Weapons/Ares.png', 0),
        'Bucky': cv2.imread('Weapons/Bucky.png', 0),
        'Bulldog': cv2.imread('Weapons/Bulldog.png', 0),
        'Classic': cv2.imread('Weapons/Classic.png', 0),
        'Frenzy': cv2.imread('Weapons/Frenzy.png', 0),
        'Ghost': cv2.imread('Weapons/Ghost.png', 0),
        'Guardian': cv2.imread('Weapons/Guardian.png', 0),
        'Judge': cv2.imread('Weapons/Judge.png', 0),
        'Knife': cv2.imread('Weapons/Knife.png', 0),
        'Marshal': cv2.imread('Weapons/Marshal.png', 0),
        'None': cv2.imread('Weapons/None.png', 0),
        'Odin': cv2.imread('Weapons/Odin.png', 0),
        'Operator': cv2.imread('Weapons/Operator.png', 0),
        'Outlaw': cv2.imread('Weapons/Outlaw.png', 0),
        'Phantom': cv2.imread('Weapons/Phantom.png', 0),
        'Vandal': cv2.imread('Weapons/Vandal.png', 0),
        'Sheriff': cv2.imread('Weapons/Sheriff.png', 0),
        'Shorty': cv2.imread('Weapons/Shorty.png', 0),
        'Spectre': cv2.imread('Weapons/Spectre.png', 0),
        'Stinger': cv2.imread('Weapons/Stinger.png', 0),
    }
    shield_templates = {
        'Shield Half': cv2.imread('Weapons/Shield Half.png', 0),
        'Shield Full': cv2.imread('Weapons/Shield Full.png', 0),
    }
    return weapon_templates, shield_templates

weapon_templates, shield_templates = load_templates()
weapon_threshold = 0.8
shield_threshold = 0.43
regions = [
    {'top': 338, 'left': 643, 'width': 170, 'height': 168},    {'top': 568, 'left': 643, 'width': 170, 'height': 168},
    {'top': 338, 'left': 1053, 'width': 118, 'height': 34},    {'top': 372, 'left': 1053, 'width': 118, 'height': 34},
    {'top': 406, 'left': 1053, 'width': 118, 'height': 34},    {'top': 440, 'left': 1053, 'width': 118, 'height': 34},
    {'top': 474, 'left': 1053, 'width': 118, 'height': 34},    {'top': 568, 'left': 1053, 'width': 118, 'height': 34},
    {'top': 602, 'left': 1053, 'width': 118, 'height': 34},    {'top': 636, 'left': 1053, 'width': 118, 'height': 34},
    {'top': 670, 'left': 1053, 'width': 118, 'height': 34},    {'top': 704, 'left': 1053, 'width': 118, 'height': 34},
    {'top': 338, 'left': 1171, 'width': 35, 'height': 34},    {'top': 372, 'left': 1171, 'width': 35, 'height': 34},
    {'top': 406, 'left': 1171, 'width': 35, 'height': 34},    {'top': 440, 'left': 1171, 'width': 35, 'height': 34},
    {'top': 474, 'left': 1171, 'width': 35, 'height': 34},    {'top': 568, 'left': 1171, 'width': 35, 'height': 34},
    {'top': 602, 'left': 1171, 'width': 35, 'height': 34},    {'top': 636, 'left': 1171, 'width': 35, 'height': 34},
    {'top': 670, 'left': 1171, 'width': 35, 'height': 34},    {'top': 704, 'left': 1171, 'width': 35, 'height': 34},
]

def resize_template(template, region_width, region_height):
    return cv2.resize(template, (region_width, region_height), interpolation=cv2.INTER_AREA)

def template_matching(region_img, templates, threshold=0.8):
    region_gray = preprocess_image(region_img)

    matched_templates = []
    region_height, region_width = region_gray.shape

    for name, template in templates.items():
        if template is None:
            continue

        resized_template = resize_template(template, region_width, region_height)
        result = cv2.matchTemplate(region_gray, resized_template, cv2.TM_CCOEFF_NORMED)
        max_val = np.max(result)

        if max_val >= threshold:
            matched_templates.append(name)

    return matched_templates if matched_templates else ["None"]

def extract_names(img_region):
    preprocessed_img = preprocess_image(img_region)
    result = ocr.ocr(preprocessed_img, cls=False)
    extracted_text = [item[1][0] for line in result for item in line]
    return extracted_text if extracted_text else ["Unknown"]

def save_to_excel(data, filename="scoreboard.xlsx", sheet_name="loadout"):
    try:
        workbook = openpyxl.load_workbook(filename)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet.delete_rows(1, sheet.max_row + 1)
        else:
            sheet = workbook.create_sheet(sheet_name)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

    headers = ['Name', 'Weapon', 'Shield']
    sheet.append(headers)

    for row in data:
        sheet.append(row)

    workbook.save(filename)
    print(f"Data saved to {filename}")

extracted_data = []

with mss.mss() as sct:
    names = []
    weapons = []
    shields = []

    for i, region in enumerate(regions):
        screenshot = sct.grab(region)
        img = np.array(screenshot)
        img = cv2.cvtColor(img, cv2.COLOR_BGRA2BGR)

        if i in [0, 1]:
            extracted_names = extract_names(img)
            names.extend(extracted_names)
        elif 2 <= i <= 11:
            matched_weapon = template_matching(img, weapon_templates, threshold=weapon_threshold)
            weapons.append(matched_weapon[0])
        elif 12 <= i <= 21:
            matched_shield = template_matching(img, shield_templates, threshold=shield_threshold)
            shields.append(matched_shield[0])

    for idx in range(len(names)):
        name = names[idx] if idx < len(names) else "Unknown"
        weapon = weapons[idx] if idx < len(weapons) else "None"
        shield = shields[idx] if idx < len(shields) else "None"
        extracted_data.append([name, weapon, shield])

    save_to_excel(extracted_data)

