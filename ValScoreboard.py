import cv2
import numpy as np
import mss
from paddleocr import PaddleOCR
import re
import openpyxl
import time

ocr = PaddleOCR(use_angle_cls=False, lang='en', use_gpu=True)
regions = [
    {'top': 338, 'left': 643, 'width': 170, 'height': 168},
    {'top': 568, 'left': 643, 'width': 170, 'height': 168},
    {'top': 338, 'left': 813, 'width': 120, 'height': 168},
    {'top': 568, 'left': 813, 'width': 120, 'height': 168},
    {'top': 338, 'left': 933, 'width': 40, 'height': 168},
    {'top': 568, 'left': 933, 'width': 40, 'height': 168},
    {'top': 338, 'left': 973, 'width': 40, 'height': 168},
    {'top': 568, 'left': 973, 'width': 40, 'height': 168},
    {'top': 338, 'left': 1013, 'width': 40, 'height': 168},
    {'top': 568, 'left': 1013, 'width': 40, 'height': 168},
    {'top': 338, 'left': 1203, 'width': 90, 'height': 168},
    {'top': 568, 'left': 1203, 'width': 90, 'height': 168},
]

def clean_ocr_data(ocr_text):
    clean_data = []
    for line in ocr_text:
        cleaned_line = re.sub(r'[^\w\s/]', '', line)  # Remove non-alphanumeric characters
        clean_data.append(cleaned_line.strip())
    return clean_data

def process_ocr_results(names, ultimates, kills, deaths, assists, credits):
    data = []
    for i in range(len(names)):
        row = [
            names[i] if i < len(names) else '',
            ultimates[i] if i < len(ultimates) else '',
            kills[i] if i < len(kills) else '',
            deaths[i] if i < len(deaths) else '',
            assists[i] if i < len(assists) else '',
            credits[i] if i < len(credits) else ''
        ]
        data.append(row)
    return data

import openpyxl

def save_to_excel(data, filename="scoreboard.xlsx", sheet_name="live"):
    try:
        workbook = openpyxl.load_workbook(filename)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet.delete_rows(1, sheet.max_row)
        else:
            sheet = workbook.create_sheet(sheet_name)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

    headers = ['Player Name', 'Ultimate', 'Kills', 'Deaths', 'Assists', 'Credits']
    sheet.append(headers)

    for row in data:
        sheet.append(row)

    workbook.save(filename)
    workbook.close()
    print(f"Data saved to {filename}")

with mss.mss() as sct:
    region_results = {
        'names': [],
        'ultimates': [],
        'kills': [],
        'deaths': [],
        'assists': [],
        'credits': [],
    }

    for i, region in enumerate(regions):
        screenshot = sct.grab(region)
        img = np.array(screenshot)
        img = cv2.cvtColor(img, cv2.COLOR_BGRA2BGR)

        scaled_img = cv2.resize(img, None, fx=1.7, fy=1.7, interpolation=cv2.INTER_CUBIC)

        gray_img = cv2.cvtColor(scaled_img, cv2.COLOR_BGR2GRAY)
        _, thresh_img = cv2.threshold(gray_img, 135, 255, cv2.THRESH_BINARY_INV)

        result = ocr.ocr(thresh_img, cls=False)

        if result is not None and len(result) > 0:
            extracted_text = [item[1][0] for line in result if line is not None for item in line if item is not None]
            cleaned_result = clean_ocr_data(extracted_text)
        else:
            cleaned_result = []

        if i == 0 or i == 1:
            region_results['names'].extend(cleaned_result)
        elif i == 2 or i == 3:
            region_results['ultimates'].extend(cleaned_result)
        elif i == 4 or i == 5:
            region_results['kills'].extend(cleaned_result)
        elif i == 6 or i == 7:
            region_results['deaths'].extend(cleaned_result)
        elif i == 8 or i == 9:
            region_results['assists'].extend(cleaned_result)
        elif i == 10 or i == 11:
            region_results['credits'].extend(cleaned_result)

processed_data = process_ocr_results(
    region_results['names'],
    region_results['ultimates'],
    region_results['kills'],
    region_results['deaths'],
    region_results['assists'],
    region_results['credits'],
)

save_to_excel(processed_data)